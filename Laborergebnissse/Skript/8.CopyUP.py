# -*- coding: utf-8 -*-
"""
Created on Tue Jul  4 16:00:05 2023

@author: zeh
"""

import openpyxl

# Load the source workbook
source_workbook = openpyxl.load_workbook('C:/Users/zeh/Desktop/Laborergebnissse/Vorlage_UP_Laborergebnisse.xlsx')
source_sheet = source_workbook.active

# Load the destination workbook
destination_workbook = openpyxl.load_workbook('C:/Users/zeh/Desktop/Laborergebnissse/Transferred.xlsx')
destination_sheet = destination_workbook['Transferred']

# Find the last row in the destination sheet
last_row = destination_sheet.max_row + 1

# Iterate over rows in the source sheet (starting from row 3)
for row in source_sheet.iter_rows(min_row=3, values_only=True):
    # Extract columns A to Z from the row
    row_data = row[:21]

    # Copy the row data to the destination sheet
    destination_sheet.append(row_data)

# Save the changes to the destination workbook
destination_workbook.save('C:/Users/zeh/Desktop/Laborergebnissse/Transferred.xlsx')

