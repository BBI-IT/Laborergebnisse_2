import openpyxl

# Load the consolidated workbook
consolidated_workbook = openpyxl.load_workbook('C:/Users/zeh/Desktop/Laborergebnissse/Consolidated.xlsx')

# Create a new workbook for the transferred data
transferred_workbook = openpyxl.Workbook()
transferred_sheet = transferred_workbook.active

# Specify the range to be transferred (columns D to K, rows 6 to 26)
transfer_range = 'D6:K26'

# Iterate through each sheet in the consolidated workbook
for sheet_name in consolidated_workbook.sheetnames:
    sheet = consolidated_workbook[sheet_name]

    # Iterate through each column in the transfer range
    for column in sheet.iter_cols(min_row=6, max_row=26, min_col=4, max_col=11):
        # Copy the column values into a row in the transferred sheet
        row_values = [cell.value for cell in column]
        transferred_sheet.append(row_values)

# Save the transferred workbook with the new name
transferred_workbook.save('C:/Users/zeh/Desktop/Laborergebnissse/Transferred.xlsx')

# Append the transferred data to the Consolidated.xlsx file
consolidated_workbook_transferred = openpyxl.load_workbook('C:/Users/zeh/Desktop/Laborergebnissse/Consolidated.xlsx')
consolidated_workbook_transferred.create_sheet('Transferred')
transferred_sheet = transferred_workbook.active

for row in transferred_sheet.iter_rows(values_only=True):
    consolidated_workbook_transferred['Transferred'].append(row)

consolidated_workbook_transferred.save('C:/Users/zeh/Desktop/Laborergebnissse/Transferred.xlsx')
