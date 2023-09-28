# -*- coding: utf-8 -*-
"""
Created on Tue Jun 20 17:36:44 2023

@author: zeh
"""

import openpyxl

# Load the workbook
workbook = openpyxl.load_workbook(r'C:\Users\zeh\Desktop\Laborergebnissse\AnlagenNummer.xlsx')
#Load the attachment Number; To run: 1. write "#" in front of row above; 2. Erase "#" in row below
#workbook = openpyxl.load_workbook(r'C:\Users\zeh\Desktop\Laborergebnissse\KV_Nummer.xlsx')

# Select the "Transferred" spreadsheet
sheet = workbook['Transferred']

# Iterate through each cell in the sheet
for row in sheet.iter_rows():
    for cell in row:
        if cell.value is not None and str(cell.value).startswith('B BW'):
            parts = str(cell.value).split(' ')
            if len(parts) == 3 and '.' in parts[2]:
                number_parts = parts[2].split('.')
                if len(number_parts) == 2:
                    n = int(number_parts[0])
                    if n < 10:
                        number_parts[0] = f'0{n}'
                        parts[2] = '.'.join(number_parts)
                        new_value = ' '.join(parts)
                        cell.value = new_value

# Save the modified workbook
workbook.save(r'C:\Users\zeh\Desktop\Laborergebnissse\AnlagenNummer.xlsx')
# Save the modified workbook
#workbook.save(r'C:\Users\zeh\Desktop\Laborergebnissse\KV_NummerI.xlsx')