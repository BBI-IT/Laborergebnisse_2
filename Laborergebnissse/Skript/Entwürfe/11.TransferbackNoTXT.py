# -*- coding: utf-8 -*-
"""
Created on Tue Jul  4 11:29:43 2023

@author: zeh
"""

import openpyxl

# Define the file paths and names
source_file = r'C:\Users\zeh\Desktop\Laborergebnissse\AnlagenNummer.xlsx'
destination_file = r'C:\Users\zeh\Desktop\Laborergebnissse\TestTransfer.xlsx'

# Load the source workbook
source_workbook = openpyxl.load_workbook(source_file)

# Get the "Transferred" sheet from the source workbook
source_sheet = source_workbook['Transferred']

# Get the list of destination sheets
dest_sheets = [source_workbook[sheet_name] for sheet_name in source_workbook.sheetnames if sheet_name.startswith('Anlage_4')]

# Transfer rows to columns across the destination sheets
row_index = 1  # Start with the first row
dest_sheet_index = 0  # Start with the first destination sheet
dest_col_index = 4  # Start with column D in the destination sheet
last_copied_sheet = None
last_copied_column = None

while row_index <= source_sheet.max_row:
    for col_index, cell in enumerate(source_sheet.iter_cols(min_row=row_index, max_row=row_index, values_only=True)):
        dest_sheet = dest_sheets[dest_sheet_index]
        dest_cell = dest_sheet.cell(row=6 + col_index, column=dest_col_index)
        dest_cell.value = cell[0]
        dest_cell.number_format = source_sheet.cell(row=row_index, column=1).number_format

        last_copied_sheet = dest_sheet.title
        last_copied_column = dest_cell.column_letter

    row_index += 1  # Move to the next row
    dest_col_index += 1  # Move to the next column in the current destination sheet

    if dest_col_index > 11 or row_index > source_sheet.max_row:
        # Move to the next destination sheet when all columns are filled or reached the last row
        dest_sheet_index += 1
        dest_col_index = 4  # Start with column D in the new destination sheet

# Save the modified source workbook as the destination workbook
source_workbook.save(destination_file)

# Remove the remaining destination sheets and clear their data
for i in range(dest_sheet_index + 1, len(dest_sheets)):
    dest_sheet = dest_sheets[i]
    source_workbook.remove(dest_sheet)

# Save the modified source workbook as the destination workbook
source_workbook.save(destination_file)

# Print the last transferred spreadsheet and column
print(f"Last transferred spreadsheet: {last_copied_sheet}")
print(f"Last transferred column: {last_copied_column}")
