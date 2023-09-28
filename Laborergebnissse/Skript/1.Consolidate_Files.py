import os
import openpyxl
from copy import copy

# Path of the folder containing the spreadsheets
subfolder_path = 'C:/Users/zeh/Desktop/Laborergebnissse/Eintragen'

# Create a new workbook
consolidated_workbook = openpyxl.Workbook()

# Iterate over the files in the subfolder
for root, dirs, files in os.walk(subfolder_path):
    for file in files:
        # Check if the file is an Excel file
        if file.endswith('.xlsx'):
            # Open the spreadsheet
            spreadsheet_path = os.path.join(root, file)
            spreadsheet_workbook = openpyxl.load_workbook(spreadsheet_path, read_only=True)

            # Iterate over the sheets in the spreadsheet
            for sheet_name in spreadsheet_workbook.sheetnames:
                # Get the sheet from the spreadsheet
                source_sheet = spreadsheet_workbook[sheet_name]

                # Create a new sheet in the consolidated workbook
                consolidated_sheet = consolidated_workbook.create_sheet(title=sheet_name)

                # Iterate over the rows in the source sheet
                for source_row in source_sheet.iter_rows():
                    # Create a new row in the consolidated sheet
                    consolidated_row = []
                    for source_cell in source_row:
                        # Check if the source cell is empty
                        if source_cell.value is None:
                            destination_cell = openpyxl.cell.cell.Cell(consolidated_sheet)
                        else:
                            # Create a new cell in the consolidated row
                            destination_cell = openpyxl.cell.cell.Cell(consolidated_sheet, column=source_cell.column, row=source_cell.row)
                            destination_cell.value = source_cell.value

                            # Copy the formatting from the source cell to the destination cell
                            if source_cell.has_style:
                                destination_cell.font = copy(source_cell.font)
                                destination_cell.border = copy(source_cell.border)
                                destination_cell.fill = copy(source_cell.fill)
                                destination_cell.number_format = copy(source_cell.number_format)
                                destination_cell.protection = copy(source_cell.protection)
                                destination_cell.alignment = copy(source_cell.alignment)

                        consolidated_row.append(destination_cell)

                    # Append the consolidated row to the consolidated sheet
                    consolidated_sheet.append(consolidated_row)

# Remove the default sheet created by openpyxl
consolidated_workbook.remove(consolidated_workbook["Sheet"])

# Save the consolidated workbook
consolidated_workbook.save('C:/Users/zeh/Desktop/Laborergebnissse/Consolidated.xlsx')
