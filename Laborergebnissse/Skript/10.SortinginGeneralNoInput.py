import openpyxl

# Load the workbook
workbook = openpyxl.load_workbook(r'C:\Users\zeh\Desktop\Laborergebnissse\AnlagenNummer.xlsx')

# Select the "Transferred" sheet
sheet = workbook['Transferred']

# Define the sorting key function
def get_sort_key(row):
    name = row[0]

    # Define the sorting priorities
    sorting_priorities = {
        "BS": 1,
        "BS AS3": 2,
        "B BW": 3,
        "B": 4
    }

    # Extract the prefix and numbers from the name
    prefix, numbers = name.split(maxsplit=1)

    # Assign a sorting priority based on the prefix
    prefix_order = sorting_priorities.get(prefix, float('inf'))

    return prefix_order, numbers

# Get all rows (including the first row) as a list
rows = list(sheet.iter_rows(values_only=True))

# Sort the rows based on the sorting key
sorted_rows = sorted(rows, key=get_sort_key)

# Clear the existing data in the sheet
sheet.delete_rows(1, sheet.max_row)

# Write the sorted rows back to the sheet
for row in sorted_rows:
    sheet.append(row)

# Save the modified workbook
workbook.save(r'C:\Users\zeh\Desktop\Laborergebnissse\AnlagenNummer.xlsx')
