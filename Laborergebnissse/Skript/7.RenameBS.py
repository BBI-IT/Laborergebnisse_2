# -*- coding: utf-8 -*-
"""
Created on Tue Jun 20 17:57:12 2023

@author: zeh
"""

import openpyxl

# Load the workbook
workbook = openpyxl.load_workbook(r'C:\Users\zeh\Desktop\Laborergebnissse\AnlagenNummer.xlsx')
#Load the attachment Number; To run: 1. write "#" in front of row above; 2. Erase "#" in row below
#workbook = openpyxl.load_workbook(r'C:\Users\zeh\Desktop\Laborergebnissse\KV_NummerI.xlsx')

# Select the "Transferred" spreadsheet
sheet = workbook['Transferred']

# Iterate through each cell in the sheet
for row in sheet.iter_rows():
    for cell in row:
        if cell.value is not None and str(cell.value).startswith('BS '):
            parts = str(cell.value).split(' ')
            if len(parts) == 2 and '/' in parts[1]:
                number_parts = parts[1].split('/')
                if len(number_parts) == 2:
                    try:
                        number = int(number_parts[0])
                        if number < 10:
                            number_parts[0] = f'00{number}'
                        elif number < 100:
                            number_parts[0] = f'0{number}'
                        parts[1] = '/'.join(number_parts)
                        new_value = ' '.join(parts)
                        cell.value = new_value
                    except ValueError:
                        pass

# Save the modified workbook
workbook.save(r'C:\Users\zeh\Desktop\Laborergebnissse\AnlagenNummer.xlsx')

# Save the modified workbook
#workbook.save(r'C:\Users\zeh\Desktop\Laborergebnissse\KV_NummerII.xlsx')
