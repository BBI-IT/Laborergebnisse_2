# -*- coding: utf-8 -*-
"""
Created on Tue Jul  4 17:02:36 2023

@author: zeh
"""

import openpyxl

# Load the KV_NummerII.xlsx file
kv_workbook = openpyxl.load_workbook('C:/Users/zeh/Desktop/Laborergebnissse/Vorlage_KVNummer.xlsx')

# Load the AnlagenNummer.xlsx file
anlagen_workbook = openpyxl.load_workbook('C:/Users/zeh/Desktop/Laborergebnissse/AnlagenNummer.xlsx')

# Get the "Transferred" sheet from the AnlagenNummer.xlsx file
transferred_sheet = anlagen_workbook['Transferred']

# Iterate over each row in the KV_NummerII.xlsx file
for row in kv_workbook.active.iter_rows(min_row=4, values_only=True):
    name = row[0]  # Name from column A
    value = row[4]  # Value from column D

    if value:
        # Search for the name in column A of the "Transferred" sheet
        for row_index, cell in enumerate(transferred_sheet['A'], start=1):
            if cell.value == name:
                # Write the value to column O of the corresponding row
                transferred_sheet.cell(row=row_index, column=15).value = value
                break

# Save the changes to the AnlagenNummer.xlsx file
anlagen_workbook.save('C:/Users/zeh/Desktop/Laborergebnissse/AnlagenNummer.xlsx')
