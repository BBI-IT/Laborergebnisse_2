# -*- coding: utf-8 -*-
"""
Created on Mon Jul  3 16:45:04 2023

@author: zeh
"""

import openpyxl
import re

# Load the workbook
workbook = openpyxl.load_workbook(r'C:\Users\zeh\Desktop\Laborergebnissse\AnlagenNummer.xlsx')

# Select the "Transferred" sheet
sheet = workbook['Transferred']

# Define the sorting key function
def get_sort_key(row):
    name = row[0]

    # Extract the prefix (Roman numeral) and numbers from the name
    match = re.search(r'^([IVXLCDM]+)\s(.+)', name)
    if match:
        prefix = match.group(1)
        numbers = match.group(2)
    else:
        prefix = ''
        numbers = name

    # Convert Roman numeral to a numeric value for sorting
    roman_values = {'I': 1, 'V': 5, 'X': 10, 'L': 50, 'C': 100, 'D': 500, 'M': 1000}
    roman_numeral = sum(roman_values[char] for char in prefix if char in roman_values)

    return roman_numeral, numbers

# Get all rows (including the first row) as a list
rows = list(sheet.iter_rows(values_only=True))

# Sort the rows based on the sorting key
sorted_rows = sorted(rows, key=get_sort_key)

# Clear the existing data in the sheet
sheet.delete_rows(1, sheet.max_row)

# Write the sorted rows back to the sheet
for row in sorted_rows:
    sheet.append(row)

# Save the modified workbook
workbook.save(r'C:\Users\zeh\Desktop\Laborergebnissse\AnlagenNummer.xlsx')

